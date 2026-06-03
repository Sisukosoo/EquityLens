"""OpenPyXL Excel report generation for CAPM, WACC, DCF, and validation."""

from __future__ import annotations

from io import BytesIO
import shutil
import subprocess
import tempfile
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import DATA_SOURCES, DISCLAIMER, current_timestamp
from utils.logger import log_event
from utils.sanity_checks import build_excel_sanity_checks, runtime_sanity_checks_for_excel


DARK_BLUE = "16213E"
WHITE = "FFFFFF"
GOLD = "F0A500"
LIGHT_BLUE = "D9EAF7"
LIGHT_GREEN = "DDEFD8"
LIGHT_YELLOW = "FFF2CC"
LIGHT_GRAY = "E7E6E6"
STATUS_GREEN = "C6EFCE"
STATUS_GREEN_TEXT = "006100"
STATUS_YELLOW = "FFEB9C"
STATUS_YELLOW_TEXT = "9C6500"
STATUS_RED = "FFC7CE"
STATUS_RED_TEXT = "9C0006"
THIN_BORDER = Border(
    left=Side(style="thin", color="A6A6A6"),
    right=Side(style="thin", color="A6A6A6"),
    top=Side(style="thin", color="A6A6A6"),
    bottom=Side(style="thin", color="A6A6A6"),
)


def build_valuation_excel_report(
    data: dict,
    income_metrics: pd.DataFrame,
    balance_metrics: pd.DataFrame,
    cash_flow_metrics: pd.DataFrame,
    beta_match: Any,
    valuation: dict[str, Any],
    validation_result: dict[str, Any] | None,
    sanity_warnings: list[dict[str, str]],
) -> bytes:
    """
    Build a complete CAPM/WACC/DCF Excel workbook.

    Formula: workbook embeds CAPM, Hamada re-levering, WACC, FCFF DCF, and validation formulas.
    Source: yfinance, Damodaran NYU Stern datasets, and valuation formulas in utils.valuation.
    Example: build_valuation_excel_report(data, statements, beta_match, valuation, validation, warnings).
    Required inputs: company data, calculated statements, matched Damodaran row, valuation dict.
    Limitation: forecasts are editable templates; automated assumptions require analyst review.
    """
    workbook = Workbook()
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    default = workbook.active
    workbook.remove(default)

    sheets = {
        "Summary": workbook.create_sheet("Summary"),
        "Beta (Damodaran)": workbook.create_sheet("Beta (Damodaran)"),
        "CAPM": workbook.create_sheet("CAPM"),
        "WACC": workbook.create_sheet("WACC"),
        "DCF": workbook.create_sheet("DCF"),
        "Validation": workbook.create_sheet("Validation"),
        "Raw Data": workbook.create_sheet("Raw Data"),
    }

    _build_summary(sheets["Summary"], data, beta_match, valuation, sanity_warnings)
    _build_beta_sheet(sheets["Beta (Damodaran)"], data, beta_match, valuation)
    _build_capm_sheet(sheets["CAPM"], valuation)
    _build_wacc_sheet(sheets["WACC"], valuation)
    _build_dcf_sheet(sheets["DCF"], income_metrics, cash_flow_metrics, valuation)
    _build_validation_sheet(sheets["Validation"], validation_result, sanity_warnings, valuation, income_metrics, data)
    _build_raw_data_sheet(sheets["Raw Data"], data, income_metrics, balance_metrics, cash_flow_metrics, beta_match)

    for sheet in sheets.values():
        _autofit(sheet)
        sheet.freeze_panes = "A2"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return _recalculate_workbook_formulas(output.getvalue())


def _find_libreoffice_executable() -> str | None:
    """Locate a LibreOffice executable for headless workbook recalculation."""
    for command in ("libreoffice", "soffice", "soffice.exe", "soffice.com"):
        path = shutil.which(command)
        if path:
            return path
    for candidate in (
        Path(r"C:\Program Files\LibreOffice\program\soffice.exe"),
        Path(r"C:\Program Files\LibreOffice\program\soffice.com"),
        Path(r"C:\Program Files (x86)\LibreOffice\program\soffice.exe"),
        Path(r"C:\Program Files (x86)\LibreOffice\program\soffice.com"),
    ):
        if candidate.exists():
            return str(candidate)
    return None


def _recalculate_workbook_formulas(workbook_bytes: bytes) -> bytes:
    """Populate cached Excel formula values with LibreOffice when available."""
    executable = _find_libreoffice_executable()
    if not executable:
        log_event("LibreOffice not found; Excel formulas will recalculate when the workbook is opened.", "excel_warning")
        return workbook_bytes
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            input_path = temp_path / "valuation_input.xlsx"
            output_dir = temp_path / "recalculated"
            output_dir.mkdir()
            input_path.write_bytes(workbook_bytes)
            subprocess.run(
                [
                    executable,
                    "--headless",
                    "--calc",
                    "--convert-to",
                    "xlsx",
                    "--outdir",
                    str(output_dir),
                    str(input_path),
                ],
                check=True,
                timeout=60,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
            )
            recalculated = output_dir / input_path.name
            if not recalculated.exists():
                log_event("LibreOffice recalc finished but no recalculated workbook was produced.", "excel_warning")
                return workbook_bytes
            return recalculated.read_bytes()
    except Exception as exc:
        log_event(f"LibreOffice formula recalc failed: {exc}", "excel_warning")
        return workbook_bytes


def _title(sheet, title: str) -> None:
    """Apply a worksheet title band on the first row."""
    _title_at(sheet, title, 1)


def _title_at(sheet, title: str, row: int) -> None:
    """
    Apply a worksheet title band.

    Formula: not applicable.
    Source: internal workbook styling.
    Example: _title(sheet, "CAPM").
    Required inputs: worksheet and title.
    Limitation: assumes title spans columns A:F.
    """
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    cell = sheet.cell(row, 1)
    cell.value = title
    cell.font = Font(bold=True, size=14, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
    cell.alignment = Alignment(horizontal="center")


def _header(row_range) -> None:
    """
    Style a header row.

    Formula: not applicable.
    Source: internal workbook styling.
    Example: _header(sheet["A3:F3"]).
    Required inputs: openpyxl row range.
    Limitation: only styles passed cells.
    """
    for cell in row_range[0]:
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
        cell.border = THIN_BORDER


def _subheader(cell) -> None:
    """
    Style a section subheader.

    Formula: not applicable.
    Source: internal workbook styling.
    Example: _subheader(sheet["A5"]).
    Required inputs: openpyxl cell.
    Limitation: single-cell style only.
    """
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)


def _input(cell) -> None:
    """
    Style a user-editable input cell.

    Formula: not applicable.
    Source: user requirement for yellow editable cells.
    Example: _input(sheet["B8"]).
    Required inputs: openpyxl cell.
    Limitation: does not lock or protect workbook.
    """
    cell.fill = PatternFill("solid", fgColor=LIGHT_YELLOW)
    cell.border = THIN_BORDER


def _formula(cell) -> None:
    """
    Style a formula cell.

    Formula: not applicable.
    Source: user requirement for light-blue formulas.
    Example: _formula(sheet["B12"]).
    Required inputs: openpyxl cell.
    Limitation: Excel recalculates formulas when opened.
    """
    cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    cell.border = THIN_BORDER


def _result(cell) -> None:
    """
    Style a key output cell.

    Formula: not applicable.
    Source: user requirement for green result cells.
    Example: _result(sheet["B14"]).
    Required inputs: openpyxl cell.
    Limitation: styling only.
    """
    cell.fill = PatternFill("solid", fgColor=LIGHT_GREEN)
    cell.font = Font(bold=True)
    cell.border = THIN_BORDER


def _pct(cell) -> None:
    """
    Apply percentage number format.

    Formula: not applicable.
    Source: Excel number-format convention.
    Example: _pct(sheet["B2"]).
    Required inputs: openpyxl cell.
    Limitation: assumes decimal input.
    """
    cell.number_format = "0.00%"


def _multiple(cell) -> None:
    """
    Apply multiple number format.

    Formula: not applicable.
    Source: Excel number-format convention.
    Example: _multiple(sheet["B2"]).
    Required inputs: openpyxl cell.
    Limitation: assumes numeric input.
    """
    cell.number_format = '0.00"x"'


def _money(cell, currency: str) -> None:
    """
    Apply millions currency number format.

    Formula: not applicable.
    Source: Excel number-format convention.
    Example: _money(cell, "EUR") -> #,##0.0 "M EUR".
    Required inputs: cell and currency code.
    Limitation: no FX conversion is performed.
    """
    cell.number_format = f'#,##0.0 "M {currency}"'


def _price(cell, currency: str) -> None:
    """
    Apply single-share price currency formatting.

    Formula: not applicable.
    Source: Excel number-format convention.
    Example: _price(cell, "USD") -> #,##0.00 "USD".
    Required inputs: cell and currency code.
    Limitation: no FX conversion is performed.
    """
    cell.number_format = f'#,##0.00 "{currency}"'


def _decimal_to_percentage_points(value: float | int | None) -> float | None:
    """Convert a decimal difference to percentage points for display."""
    if value is None or pd.isna(value):
        return None
    return float(value) * 100


def _tier_status_text(tier: dict[str, Any], critical_sanity_warning: bool = False) -> str:
    """Return report status text with a non-empty reason for accepted tiers."""
    if tier.get("tier") == 5 and tier.get("status") == "REFERENCE":
        return "REFERENCE - tangible book value shown as conservative floor"
    if tier.get("selected"):
        if critical_sanity_warning:
            return "Computed but flagged - critical sanity check applies. See Validation tab."
        reason = tier.get("acceptance_reason") or tier.get("selection_reason") or "within tier sanity limits"
        return f"Used as primary result - {reason}"
    status = tier.get("status") or "UNKNOWN"
    if status == "ACCEPTED":
        reason = tier.get("acceptance_reason") or "within tier sanity limits"
        return f"ACCEPTED - {reason}"
    reason = tier.get("rejection_reason") or tier.get("skip_message") or "no reason provided"
    return f"{status} - {reason}"


def _build_summary(
    sheet,
    data: dict,
    beta_match: Any,
    valuation: dict[str, Any],
    sanity_warnings: list[dict[str, str]] | None = None,
) -> None:
    """
    Build the Summary worksheet.

    Formula: pulls outputs from the valuation package and links to detail tabs.
    Source: yfinance + Damodaran valuation result.
    Example: first worksheet users see.
    Required inputs: company data, beta match, valuation dict.
    Limitation: summary values are snapshots at generation time.
    """
    has_critical_sanity_warning = _has_critical_sanity_warning(sanity_warnings)
    if has_critical_sanity_warning:
        _critical_summary_warning(sheet)
        _title_at(sheet, "Valuation Summary", 2)
    else:
        _title(sheet, "Valuation Summary")
    info = data.get("info", {})
    currency = valuation.get("currency", "")
    selected = valuation.get("selected_dcf_tier") or {}
    selected_dcf = valuation.get("dcf") or {}
    rows = [
        ("Company", info.get("longName") or info.get("shortName")),
        ("Ticker", data.get("ticker")),
        ("Industry", info.get("industry")),
        ("Date", current_timestamp()),
        ("Reported currency", valuation.get("currency")),
        ("Beta", valuation.get("levered_beta")),
        ("Cost of Equity", valuation.get("cost_of_equity")),
        ("Cost of Debt", valuation.get("cost_of_debt")),
        ("Cost of Debt Estimated", "YES" if valuation.get("cost_of_debt_estimated") else "NO"),
        ("WACC", valuation.get("wacc")),
        ("Selected estimate", selected_dcf.get("implied_price")),
        ("Selected valuation tier", selected.get("method")),
        ("Current Market Price", valuation.get("current_price")),
        ("Upside / Downside", selected_dcf.get("upside")),
        ("Confidence", _summary_confidence_label(selected, has_critical_sanity_warning)),
        ("Reason", _summary_reason(selected, valuation, has_critical_sanity_warning)),
    ]
    header_row = 3 if has_critical_sanity_warning else 2
    first_data_row = header_row + 1
    sheet.append(["Field", "Value"])
    _header(sheet[f"A{header_row}:B{header_row}"])
    for row in rows:
        sheet.append(list(row))
    percent_rows = {"Cost of Equity", "Cost of Debt", "WACC", "Upside / Downside"}
    price_rows = {"Selected estimate", "Current Market Price"}
    for row_idx in range(first_data_row, first_data_row + len(rows)):
        label = sheet[f"A{row_idx}"].value
        if label in percent_rows:
            _pct(sheet[f"B{row_idx}"])
        if label == "Beta":
            _multiple(sheet[f"B{row_idx}"])
        if label in price_rows:
            _price(sheet[f"B{row_idx}"], currency)
        if label in {"WACC", "Selected estimate"}:
            _result(sheet[f"B{row_idx}"])

    tier_start = first_data_row + len(rows) + 2
    sheet[f"A{tier_start}"] = "Implied Share Price Analysis"
    _subheader(sheet[f"A{tier_start}"])
    tier_header = tier_start + 1
    sheet[f"A{tier_header}"], sheet[f"B{tier_header}"], sheet[f"C{tier_header}"], sheet[f"D{tier_header}"] = (
        "Method",
        "Implied share price",
        "Upside / downside",
        "Status",
    )
    _header(sheet[f"A{tier_header}:D{tier_header}"])
    for row_idx, tier in enumerate(valuation.get("dcf_tiers", []) or [], start=tier_header + 1):
        dcf = tier.get("dcf") or {}
        sheet[f"A{row_idx}"] = f"Tier {tier.get('tier')} - {tier.get('name')}"
        sheet[f"B{row_idx}"] = dcf.get("implied_price")
        sheet[f"C{row_idx}"] = dcf.get("upside")
        sheet[f"D{row_idx}"] = _tier_status_text(tier, has_critical_sanity_warning)
        _price(sheet[f"B{row_idx}"], currency)
        _pct(sheet[f"C{row_idx}"])
        if tier.get("selected"):
            _result(sheet[f"B{row_idx}"])
            if has_critical_sanity_warning:
                sheet[f"D{row_idx}"].fill = PatternFill("solid", fgColor=STATUS_RED)
                sheet[f"D{row_idx}"].font = Font(bold=True, color=STATUS_RED_TEXT)
            else:
                sheet[f"D{row_idx}"].fill = PatternFill("solid", fgColor=STATUS_GREEN)
                sheet[f"D{row_idx}"].font = Font(bold=True, color=STATUS_GREEN_TEXT)
        elif tier.get("status") == "REJECTED":
            sheet[f"D{row_idx}"].fill = PatternFill("solid", fgColor=STATUS_YELLOW)
            sheet[f"D{row_idx}"].font = Font(bold=True, color=STATUS_YELLOW_TEXT)

    navigation_row = tier_header + max(len(valuation.get("dcf_tiers", []) or []), 1) + 3
    sheet[f"A{navigation_row}"] = "Workbook navigation"
    sheet[f"B{navigation_row}"] = None
    _subheader(sheet[f"A{navigation_row}"])
    for idx, name in enumerate(["Beta (Damodaran)", "CAPM", "WACC", "DCF", "Validation", "Raw Data"], start=navigation_row + 1):
        sheet[f"A{idx}"] = name
        sheet[f"A{idx}"].hyperlink = f"#'{name}'!A1"
        sheet[f"A{idx}"].style = "Hyperlink"
    sheet["D3"] = "Beta source"
    _subheader(sheet["D3"])
    sheet["D4"] = f"Damodaran industry: {beta_match.matched_industry}"
    sheet["D5"] = f"Match confidence: {beta_match.confidence:.1f}%"
    sheet["D6"] = f"Updated: {beta_match.source_updated}"


def _has_critical_sanity_warning(sanity_warnings: list[dict[str, str]] | None) -> bool:
    """Return True when report generation was allowed despite a critical sanity warning."""
    return any(str(warning.get("severity", "")).lower() == "critical" for warning in sanity_warnings or [])


def _critical_summary_warning(sheet) -> None:
    """Write the top-of-summary critical warning banner."""
    sheet.merge_cells("A1:F1")
    cell = sheet["A1"]
    cell.value = (
        "⚠ CRITICAL WARNING: This valuation has critical sanity check warnings. Do not use the implied price "
        "as a fair value estimate. See Validation tab > Sanity Checks for explanation."
    )
    cell.font = Font(bold=True, color=STATUS_RED_TEXT)
    cell.fill = PatternFill("solid", fgColor=STATUS_RED)
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    cell.border = THIN_BORDER


def _summary_confidence_label(selected: dict[str, Any], critical_sanity_warning: bool) -> str:
    """Return Summary confidence text, overriding normal confidence for critical sanity warnings."""
    if critical_sanity_warning:
        return (
            "CRITICAL - DCF model not appropriate for this business type. See Validation tab. "
            "Result should not be used as a fair value estimate."
        )
    return _confidence_label(selected)


def _summary_reason(selected: dict[str, Any], valuation: dict[str, Any], critical_sanity_warning: bool) -> str:
    """Return Summary reason text, overriding normal tier explanation for critical sanity warnings."""
    if critical_sanity_warning:
        return "One or more critical sanity checks have fired. See Validation tab > Sanity Checks for details."
    return selected.get("explanation") or valuation.get("dcf_selection_reason")


def _build_beta_sheet(sheet, data: dict, beta_match: Any, valuation: dict[str, Any]) -> None:
    """
    Build the Beta (Damodaran) worksheet.

    Formula: Levered Beta = Unlevered Beta x (1 + (1 - Tax) x D/E).
    Source: Aswath Damodaran, NYU Stern industry beta tables.
    Example: shows asset beta, D/E, tax, re-levering, and yfinance reference beta.
    Required inputs: beta match and valuation dict.
    Limitation: industry average beta may not capture company-specific operating risk.
    """
    _title(sheet, "Beta (Damodaran)")
    sheet["A3"] = f"Source: Aswath Damodaran, NYU Stern - {beta_match.source_url} - Updated {beta_match.source_updated}"
    sheet["A5"], sheet["B5"] = "Company industry (yfinance)", beta_match.company_industry
    sheet["A6"], sheet["B6"] = "Matched Damodaran industry", beta_match.matched_industry
    sheet["A7"], sheet["B7"] = "Match confidence", beta_match.confidence / 100
    sheet["A8"], sheet["B8"] = "Unlevered Beta", valuation.get("unlevered_beta")
    sheet["A9"], sheet["B9"] = "Company D/E ratio", valuation.get("de_ratio")
    sheet["A10"], sheet["B10"] = "Tax rate", valuation.get("tax_rate")
    sheet["C10"] = (
        "Company effective tax rate used in relevering. Alternative methodology would use marginal tax rate "
        "(Damodaran's recommendation); results may differ for firms with significant tax shields."
    )
    sheet["A11"], sheet["B11"] = "Levered Beta", "=B8*(1+(1-B10)*B9)"
    sheet["A12"], sheet["B12"] = "yfinance beta (reference only)", valuation.get("yfinance_beta")
    sheet["C12"] = (
        "Yahoo Finance beta typically reflects 1-2 year regression against the market, which can be heavily "
        "influenced by recent market regime. The Damodaran sector-relevered beta is preferred for DCF as it "
        "represents long-run business risk and is normalized across the sector."
    )
    for row in (7, 9, 10):
        _pct(sheet[f"B{row}"])
    _formula(sheet["B11"])
    _multiple(sheet["B8"])
    _multiple(sheet["B11"])
    sheet["B11"].comment = Comment(
        "Damodaran industry beta is often more stable than a single-stock regression beta because it averages operating risk across comparable companies.",
        "Financial Analyzer",
    )
    if valuation.get("unlevered_beta_estimated"):
        sheet["C8"] = "ESTIMATED - Damodaran unlevered beta missing for matched industry; market-average 1.0 assumed."
        _input(sheet["C8"])
    if valuation.get("de_estimated"):
        sheet["C9"] = "ESTIMATED - company D/E missing, Damodaran industry average used."
        _input(sheet["C9"])
    if valuation.get("tax_estimated"):
        sheet["C10"] = (
            "ESTIMATED - company tax rate invalid/missing, Damodaran industry average used. "
            "Alternative methodology would use marginal tax rate (Damodaran's recommendation)."
        )
        _input(sheet["C10"])


def _build_capm_sheet(sheet, valuation: dict[str, Any]) -> None:
    """
    Build the CAPM worksheet.

    Formula: Re = Rf + beta x ERP.
    Source: currency-specific Rf metadata from valuation result and Damodaran recommended 5.5% ERP.
    Example: B5 formula equals B2+B3*B4.
    Required inputs: valuation dict.
    Limitation: ERP is fixed at 5.5% per prompt.
    """
    _title(sheet, "CAPM")
    rf_currency = valuation.get("risk_free_currency") or "USD"
    rf_target = valuation.get("risk_free_target_currency") or valuation.get("currency") or rf_currency
    rf_source = valuation.get("risk_free_source_detail") or valuation.get("risk_free_source") or "Risk-free rate source"
    rf_note = f"{rf_currency} risk-free rate for {rf_target} valuation; {rf_source}; as of {valuation.get('risk_free_date')}"
    rows = [
        ("Risk-free rate", valuation.get("risk_free_rate"), rf_note),
        ("Levered Beta", "='Beta (Damodaran)'!B11", "Damodaran industry beta re-levered to company D/E"),
        ("Market Risk Premium", valuation.get("market_risk_premium"), "Hardcoded source-code constant (Damodaran ERP estimate, Jan 2026); not loaded from a live source"),
        ("Cost of Equity", "=B3+B4*B5", "CAPM formula"),
    ]
    sheet.append(["Input", "Value", "Explanation"])
    _header(sheet["A2:C2"])
    for row in rows:
        sheet.append(list(row))
    for cell_ref in ("B3", "B5", "B6"):
        _pct(sheet[cell_ref])
    _multiple(sheet["B4"])
    _formula(sheet["B6"])
    _result(sheet["B6"])
    _add_sml_image(sheet, valuation)


def _add_sml_image(sheet, valuation: dict[str, Any]) -> None:
    """
    Add a Security Market Line image to the CAPM worksheet.

    Formula: Expected return = Rf + beta x ERP.
    Source: CAPM/SML convention.
    Example: plots beta from 0 to 2 and company point.
    Required inputs: valuation dict.
    Limitation: static image generated at report time.
    """
    import matplotlib.pyplot as plt

    rf = valuation.get("risk_free_rate") or 0
    erp = valuation.get("market_risk_premium") or 0.055
    beta = valuation.get("levered_beta") or 1
    xs = [0, 0.5, 1, 1.5, 2]
    ys = [rf + x * erp for x in xs]
    fig, ax = plt.subplots(figsize=(5.5, 3.2))
    ax.plot(xs, ys, color="#16213e", linewidth=2)
    ax.scatter([beta], [rf + beta * erp], color="#f0a500", s=70)
    ax.set_title("Security Market Line")
    ax.set_xlabel("Beta")
    ax.set_ylabel("Expected return")
    ax.yaxis.set_major_formatter(lambda value, _pos: f"{value:.1%}")
    ax.grid(True, alpha=0.25)
    output = BytesIO()
    fig.tight_layout()
    fig.savefig(output, format="png", dpi=160)
    plt.close(fig)
    output.seek(0)
    image = Image(output)
    image.anchor = "E3"
    sheet.add_image(image)


def _build_wacc_sheet(sheet, valuation: dict[str, Any]) -> None:
    """
    Build the WACC worksheet.

    Formula: WACC = E/V x Re + D/V x Rd x (1-T).
    Source: Damodaran WACC framework.
    Example: formulas reference CAPM and WACC component cells.
    Required inputs: valuation dict.
    Limitation: book debt is used as proxy for market debt.
    """
    currency = valuation.get("currency", "")
    _title(sheet, "WACC")
    rows = [
        ("Market cap (E)", valuation.get("market_cap"), ""),
        (
            "Total debt (book value)",
            valuation.get("total_debt"),
            "Book value used as proxy for market value of debt. Standard simplification when debt is short-term or trades close to par.",
        ),
        ("V = E + D", "=B3+B4", ""),
        ("Equity weight = E/V", "=B3/B5", ""),
        ("Debt weight = D/V", "=B4/B5", ""),
        ("Cost of equity (Re)", "='CAPM'!B6", ""),
        ("Cost of debt (Rd)", valuation.get("cost_of_debt"), ""),
        ("Tax rate (T)", valuation.get("tax_rate"), ""),
        ("After-tax cost of debt", "=B9*(1-B10)", ""),
        ("WACC", "=(B6*B8)+(B7*B9*(1-B10))", ""),
    ]
    sheet.append(["Component", "Value", "Explanation"])
    _header(sheet["A2:C2"])
    for row in rows:
        sheet.append(list(row))
    for row in (3, 4, 5):
        _money(sheet[f"B{row}"], currency)
    for row in (6, 7, 8, 9, 10, 11, 12):
        _pct(sheet[f"B{row}"])
    for row in (5, 6, 7, 9, 12):
        _formula(sheet[f"B{row}"])
    _result(sheet["B12"])
    if valuation.get("cost_of_debt_estimated"):
        sheet["C9"] = "ESTIMATED - interest expense / total debt missing or outside 1%-15%; Damodaran industry cost of debt used."
        _input(sheet["C9"])
    chart = PieChart()
    labels = Reference(sheet, min_col=1, min_row=3, max_row=4)
    data = Reference(sheet, min_col=2, min_row=3, max_row=4)
    chart.add_data(data)
    chart.set_categories(labels)
    chart.title = "Capital Structure"
    sheet.add_chart(chart, "D3")


def _build_dcf_sheet(sheet, income: pd.DataFrame, cash: pd.DataFrame, valuation: dict[str, Any]) -> None:
    """
    Build the DCF worksheet.

    Formula: FCFF = EBIT x (1-T) + D&A - CapEx - Delta WC; TV = FCF_n x (1+g)/(WACC-g).
    Source: Damodaran FCFF DCF framework.
    Example: five-year formula-driven forecast with yellow assumption cells.
    Required inputs: historical revenue/EBIT/FCF and valuation dict.
    Limitation: D&A, CapEx, and working capital are template assumptions.
    """
    currency = valuation.get("currency", "")
    _title(sheet, "DCF")
    sheet["A3"] = "Historical Data"
    _subheader(sheet["A3"])
    hist = income.tail(5).merge(cash[["year", "free_cash_flow"]], on="year", how="left")
    hist_display, excluded_periods = _filter_empty_numeric_display_rows(hist)
    header_row = 4
    if excluded_periods:
        _write_exclusion_note(sheet, header_row, excluded_periods, 4)
        header_row += 1
    for col_idx, header in enumerate(["Period", "Revenue", "EBIT", "FCF"], start=1):
        sheet.cell(header_row, col_idx).value = header
    _header(sheet[f"A{header_row}:D{header_row}"])
    for row_idx, (_, row) in enumerate(hist_display.iterrows(), start=header_row + 1):
        sheet.cell(row_idx, 1).value = row.get("period") or row.get("year")
        sheet.cell(row_idx, 2).value = row.get("revenue")
        sheet.cell(row_idx, 3).value = row.get("ebit")
        sheet.cell(row_idx, 4).value = row.get("free_cash_flow")
    for row in range(header_row + 1, header_row + 1 + len(hist_display)):
        for col in range(2, 5):
            _money(sheet.cell(row, col), currency)

    current_row = header_row + len(hist_display) + 3
    for tier in valuation.get("dcf_tiers", []) or []:
        current_row = _write_dcf_tier_section(sheet, current_row, tier, valuation, currency)
        current_row += 2

    selected = valuation.get("selected_dcf_tier") or {}
    selected_dcf = valuation.get("dcf") or {}
    sheet[f"A{current_row}"] = "Final Selected Result"
    _subheader(sheet[f"A{current_row}"])
    final_rows = [
        ("Selected tier", selected.get("method")),
        ("Selected estimate", selected_dcf.get("implied_price")),
        ("Current market price", valuation.get("current_price")),
        ("Upside / downside", selected_dcf.get("upside")),
        ("Confidence", _confidence_label(selected)),
        ("Reason", selected.get("explanation") or valuation.get("dcf_selection_reason")),
    ]
    for row_idx, (label, value) in enumerate(final_rows, start=current_row + 1):
        sheet[f"A{row_idx}"], sheet[f"B{row_idx}"] = label, value
        if label in {"Selected estimate", "Current market price"}:
            _price(sheet[f"B{row_idx}"], currency)
        elif label == "Upside / downside":
            _pct(sheet[f"B{row_idx}"])
        if label == "Selected estimate":
            _result(sheet[f"B{row_idx}"])
    reverse_start = current_row + len(final_rows) + 2
    _write_reverse_dcf_section(sheet, reverse_start, valuation)


def _write_dcf_tier_section(sheet, start_row: int, tier: dict[str, Any], valuation: dict[str, Any], currency: str) -> int:
    """Write one DCF tier's assumptions, forecast, and result to the DCF worksheet."""
    if tier.get("tier") == 4:
        return _write_multiples_tier_section(sheet, start_row, tier, currency)
    if tier.get("tier") == 5:
        return _write_tangible_book_tier_section(sheet, start_row, tier, currency)
    title = f"Tier {tier.get('tier')} Assumptions - {tier.get('name')}"
    sheet[f"A{start_row}"] = title
    _subheader(sheet[f"A{start_row}"])
    assumptions = tier.get("assumptions") or {}
    assumption_rows = [
        ("Revenue growth % per year", assumptions.get("revenue_growth"), assumptions.get("revenue_growth_source")),
        ("EBIT margin %", assumptions.get("ebit_margin"), assumptions.get("ebit_margin_source")),
        ("Tax rate", valuation.get("tax_rate"), "Company effective tax rate"),
        ("D&A % revenue", assumptions.get("depreciation_pct_revenue"), assumptions.get("depreciation_source")),
        ("CapEx % revenue", assumptions.get("capex_pct_revenue"), assumptions.get("capex_source")),
        ("Working capital % revenue", assumptions.get("working_capital_pct_revenue"), assumptions.get("working_capital_source")),
        ("Terminal growth rate", valuation.get("terminal_growth"), valuation.get("terminal_growth_source") or "Capped long-term growth assumption"),
        ("Discount rate = WACC", valuation.get("wacc"), "WACC from CAPM + cost of debt"),
    ]
    header_row = start_row + 1
    sheet[f"A{header_row}"], sheet[f"B{header_row}"], sheet[f"C{header_row}"] = "Assumption", "Value", "Source"
    _header(sheet[f"A{header_row}:C{header_row}"])
    for row_idx, row in enumerate(assumption_rows, start=header_row + 1):
        sheet[f"A{row_idx}"], sheet[f"B{row_idx}"], sheet[f"C{row_idx}"] = row
        _pct(sheet[f"B{row_idx}"])
        _input(sheet[f"B{row_idx}"])

    if tier.get("status") == "SKIPPED":
        message_row = header_row + len(assumption_rows) + 2
        sheet[f"A{message_row}"] = tier.get("skip_message") or "Tier skipped."
        sheet[f"A{message_row}"].alignment = Alignment(wrap_text=True, vertical="top")
        sheet.merge_cells(start_row=message_row, start_column=1, end_row=message_row, end_column=5)
        status_row = message_row + 2
        sheet[f"A{status_row}"], sheet[f"B{status_row}"] = "Status", _tier_status_text(tier)
        return status_row + 1

    forecast_start = header_row + len(assumption_rows) + 2
    sheet[f"A{forecast_start}"] = f"Tier {tier.get('tier')} FCF projections"
    _subheader(sheet[f"A{forecast_start}"])
    headers = ["Year", "Revenue", "EBIT", "FCF", "PV FCF"]
    for col, header in enumerate(headers, start=1):
        sheet.cell(forecast_start + 1, col).value = header
    _header(sheet[f"A{forecast_start + 1}:E{forecast_start + 1}"])
    dcf = tier.get("dcf") or {}
    for row_idx, forecast in enumerate(dcf.get("forecast", []) or [], start=forecast_start + 2):
        sheet[f"A{row_idx}"] = f"Year {forecast.get('year')}"
        sheet[f"B{row_idx}"] = forecast.get("revenue")
        sheet[f"C{row_idx}"] = forecast.get("ebit")
        sheet[f"D{row_idx}"] = forecast.get("fcf")
        sheet[f"E{row_idx}"] = forecast.get("pv_fcf")
        for col in range(2, 6):
            _money(sheet.cell(row_idx, col), currency)

    result_row = forecast_start + 8
    result_rows = [
        ("Terminal value", dcf.get("terminal_value")),
        ("PV terminal value", dcf.get("pv_terminal_value")),
        ("Enterprise value", dcf.get("enterprise_value")),
        ("Equity value", dcf.get("equity_value")),
        ("Implied share price", dcf.get("implied_price")),
        ("Upside / downside", dcf.get("upside")),
        ("Status", _tier_status_text(tier)),
    ]
    for row_idx, (label, value) in enumerate(result_rows, start=result_row):
        sheet[f"A{row_idx}"], sheet[f"B{row_idx}"] = label, value
        if label in {"Terminal value", "PV terminal value", "Enterprise value", "Equity value"}:
            _money(sheet[f"B{row_idx}"], currency)
        elif label == "Implied share price":
            _price(sheet[f"B{row_idx}"], currency)
            if tier.get("selected"):
                _result(sheet[f"B{row_idx}"])
        elif label == "Upside / downside":
            _pct(sheet[f"B{row_idx}"])
    return result_row + len(result_rows)


def _write_reverse_dcf_section(sheet, start_row: int, valuation: dict[str, Any]) -> int:
    """Write Reverse DCF diagnostics below the selected valuation result."""
    reverse = valuation.get("reverse_dcf") or {}
    sheet[f"A{start_row}"] = "Reverse DCF Analysis"
    _subheader(sheet[f"A{start_row}"])
    header_row = start_row + 1
    sheet[f"A{header_row}"], sheet[f"B{header_row}"], sheet[f"C{header_row}"] = "Metric", "Value", "Source"
    _header(sheet[f"A{header_row}:C{header_row}"])
    rows = [
        (
            "Implied growth rate from reverse DCF",
            reverse.get("implied_growth"),
            "pct",
            "Solved from current market price using Tier 1 DCF inputs",
        ),
        (
            "Tier 1 assumed growth rate",
            reverse.get("tier1_growth"),
            "pct",
            f"Same as Tier 1 DCF assumption: {reverse.get('tier1_growth_source') or 'see Tier 1 Assumptions section'}",
        ),
        (
            "Yahoo trailing revenue growth",
            reverse.get("analyst_consensus_growth"),
            "pct",
            reverse.get("analyst_consensus_source") or "N/A - yfinance revenueGrowth unavailable",
        ),
        (
            "Gap: implied vs Tier 1 assumed",
            _decimal_to_percentage_points(reverse.get("growth_gap")),
            "pp",
            "Calculated: implied minus Tier 1 assumed (in percentage points)",
        ),
        (
            "Interpretation",
            reverse.get("interpretation") or reverse.get("message"),
            "text",
            "",
        ),
    ]
    for row_idx, (label, value, value_type, source) in enumerate(rows, start=header_row + 1):
        sheet[f"A{row_idx}"], sheet[f"B{row_idx}"], sheet[f"C{row_idx}"] = label, value, source
        if value_type == "pct":
            _pct(sheet[f"B{row_idx}"])
        elif value_type == "pp":
            sheet[f"B{row_idx}"].number_format = '+0.0" pp";[Red]-0.0" pp";0.0" pp"'
    return header_row + len(rows) + 1


def _write_multiples_tier_section(sheet, start_row: int, tier: dict[str, Any], currency: str) -> int:
    """Write Tier 4 multiples valuation details to the DCF worksheet."""
    sheet[f"A{start_row}"] = "Tier 4 - Multiples-Based Valuation"
    _subheader(sheet[f"A{start_row}"])
    header_row = start_row + 1
    sheet[f"A{header_row}"], sheet[f"B{header_row}"], sheet[f"C{header_row}"], sheet[f"D{header_row}"] = (
        "Method",
        "Multiple",
        "Implied per share",
        "Source",
    )
    _header(sheet[f"A{header_row}:D{header_row}"])
    row_idx = header_row + 1
    for item in tier.get("multiples", []) or []:
        sheet[f"A{row_idx}"] = item.get("method")
        sheet[f"B{row_idx}"] = item.get("multiple")
        sheet[f"C{row_idx}"] = item.get("implied_price")
        source = item.get("source") or "N/A"
        if item.get("error"):
            source = f"{source} - {item.get('error')}"
        sheet[f"D{row_idx}"] = source
        _multiple(sheet[f"B{row_idx}"])
        _price(sheet[f"C{row_idx}"], currency)
        row_idx += 1
    dcf = tier.get("dcf") or {}
    median_source = f"Median of {tier.get('available_multiples', len(tier.get('multiples', []) or []))} methods"
    sheet[f"A{row_idx}"], sheet[f"B{row_idx}"], sheet[f"C{row_idx}"], sheet[f"D{row_idx}"] = (
        "Median (selected)",
        "-",
        dcf.get("implied_price"),
        median_source,
    )
    _price(sheet[f"C{row_idx}"], currency)
    if tier.get("selected"):
        _result(sheet[f"C{row_idx}"])
    row_idx += 1
    sheet[f"A{row_idx}"] = "Status"
    upside = dcf.get("upside")
    if tier.get("status") == "ACCEPTED":
        status_text = "ACCEPTED ✓"
        if upside is not None:
            status_text += f" ({upside:.0%} vs market, within 70% limit)"
    else:
        status_text = f"{tier.get('status')} - {tier.get('rejection_reason')}"
    detail_status = tier.get("detail_status")
    if detail_status and detail_status != "Full multiples set (3/3 available)":
        status_text = f"{status_text}; {detail_status}"
    if tier.get("selected"):
        status_text = f"{status_text}; Used as primary result"
    sheet[f"B{row_idx}"] = status_text
    row_idx += 1
    source = (tier.get("assumptions") or {}).get("source")
    sheet[f"A{row_idx}"] = "Source"
    sheet[f"B{row_idx}"] = source
    return row_idx + 1


def _write_tangible_book_tier_section(sheet, start_row: int, tier: dict[str, Any], currency: str) -> int:
    """Write Tier 5 tangible book value floor details to the DCF worksheet."""
    sheet[f"A{start_row}"] = "Tier 5 - Tangible Book Floor"
    _subheader(sheet[f"A{start_row}"])
    assumptions = tier.get("assumptions") or {}
    rows = [
        ("Tangible equity (equity - goodwill - intangibles)", assumptions.get("tangible_equity")),
        ("Shares outstanding", assumptions.get("shares_outstanding") / 1_000_000 if assumptions.get("shares_outstanding") else None),
        ("Tangible book value/share", (tier.get("dcf") or {}).get("implied_price")),
        ("Source", "(Latest balance sheet equity - goodwill - intangibles) / shares"),
        ("Status", "Used as primary result" if tier.get("selected") else "Reference"),
    ]
    for row_idx, (label, value) in enumerate(rows, start=start_row + 1):
        sheet[f"A{row_idx}"], sheet[f"B{row_idx}"] = label, value
        if label == "Tangible equity (equity - goodwill - intangibles)":
            _money(sheet[f"B{row_idx}"], currency)
        elif label == "Shares outstanding":
            sheet[f"B{row_idx}"].number_format = '#,##0.00 "M"'
        elif label == "Tangible book value/share":
            _price(sheet[f"B{row_idx}"], currency)
            if tier.get("selected"):
                _result(sheet[f"B{row_idx}"])
    return start_row + len(rows) + 1


def _confidence_label(selected: dict[str, Any]) -> str:
    """Return the user-facing confidence label for the selected DCF tier."""
    confidence = str(selected.get("confidence") or "").upper()
    tier = selected.get("tier")
    if tier == 4:
        return "LOW - multiples-based valuation; DCF model not applicable"
    if tier == 5:
        return "VERY LOW - tangible book value floor only"
    if confidence == "LOW":
        return "LOW - model relies on sector averages"
    if confidence == "MODERATE":
        return "MODERATE - model uses smoothed long-term averages"
    return "NORMAL - standard DCF assumptions"


def _build_validation_sheet(
    sheet,
    validation_result: dict[str, Any] | None,
    sanity_warnings: list[dict[str, str]],
    valuation: dict[str, Any],
    income_metrics: pd.DataFrame,
    data: dict[str, Any],
) -> None:
    """
    Build the Validation worksheet.

    Formula: difference = calculated - Damodaran benchmark.
    Source: Damodaran validation datasets and sanity-check thresholds.
    Example: WACC calculated vs industry average.
    Required inputs: validation result and sanity warnings.
    Limitation: validation depends on live Damodaran benchmark availability.
    """
    _title(sheet, "Validation")
    runtime_rows = runtime_sanity_checks_for_excel(sanity_warnings)
    analyst_rows = build_excel_sanity_checks(valuation, income_metrics, validation_result, data)
    sanity_rows = _merge_excel_sanity_rows(runtime_rows, analyst_rows)
    if not validation_result:
        sheet["A3"] = "Validation was not available."
        next_row = _write_excel_sanity_checks(sheet, 5, sanity_rows)
        _write_validation_dcf_tier_used(sheet, next_row, valuation)
        return
    sheet["A3"] = f"Source: {validation_result.get('source_url')} - Updated {validation_result.get('source_updated')}"
    sheet.append(["Metric", "Calculated", "Damodaran (industry avg)", "Difference", "Status", "Note"])
    _header(sheet["A4:F4"])
    for row in validation_result.get("rows", []):
        difference = row["Difference"]
        sheet.append([row["Metric"], row["Calculated"], row["Damodaran (industry avg)"], difference, row["Status"], row.get("Note", "")])
    for row in range(5, 5 + len(validation_result.get("rows", []))):
        for col in (2, 3):
            _pct(sheet.cell(row, col))
        sheet.cell(row, 4).number_format = '+0.00%"pp";[Red]-0.00%"pp";0.00"pp"'
        status_cell = sheet.cell(row, 5)
        if "OK" in str(status_cell.value):
            status_cell.fill = PatternFill("solid", fgColor=STATUS_GREEN)
            status_cell.font = Font(bold=True, color=STATUS_GREEN_TEXT)
        elif "Warn" in str(status_cell.value) or "Review" in str(status_cell.value):
            status_cell.fill = PatternFill("solid", fgColor=STATUS_YELLOW)
            status_cell.font = Font(bold=True, color=STATUS_YELLOW_TEXT)
        elif "Check" in str(status_cell.value) or "Critical" in str(status_cell.value):
            status_cell.fill = PatternFill("solid", fgColor=STATUS_RED)
            status_cell.font = Font(bold=True, color=STATUS_RED_TEXT)
        status_cell.border = THIN_BORDER
    data_end = 4 + len(validation_result.get("rows", []))
    if data_end >= 5:
        sheet.conditional_formatting.add(
            f"E5:E{data_end}",
            FormulaRule(formula=['ISNUMBER(SEARCH("OK",E5))'], fill=PatternFill("solid", fgColor=STATUS_GREEN), font=Font(bold=True, color=STATUS_GREEN_TEXT)),
        )
        sheet.conditional_formatting.add(
            f"E5:E{data_end}",
            FormulaRule(formula=['ISNUMBER(SEARCH("Review",E5))'], fill=PatternFill("solid", fgColor=STATUS_YELLOW), font=Font(bold=True, color=STATUS_YELLOW_TEXT)),
        )
        sheet.conditional_formatting.add(
            f"E5:E{data_end}",
            FormulaRule(formula=['OR(ISNUMBER(SEARCH("Check",E5)),ISNUMBER(SEARCH("Critical",E5)))'], fill=PatternFill("solid", fgColor=STATUS_RED), font=Font(bold=True, color=STATUS_RED_TEXT)),
        )
    start = 11
    if validation_result.get("tax_note"):
        sheet["A9"] = validation_result["tax_note"]
        sheet["A9"].font = Font(italic=True, color="666666")
        sheet.merge_cells("A9:F9")
    if validation_result.get("tax_effective_note"):
        sheet["A10"] = validation_result["tax_effective_note"]
        sheet["A10"].font = Font(italic=True, color="666666")
        sheet.merge_cells("A10:F10")
    tier_row = _write_excel_sanity_checks(sheet, start, sanity_rows)
    if valuation.get("selected_dcf_tier"):
        _write_validation_dcf_tier_used(sheet, tier_row, valuation)
        tier_row += 4

    if validation_result.get("explanations"):
        expl_row = tier_row + 1
        sheet[f"A{expl_row}"] = "Possible explanations"
        _subheader(sheet[f"A{expl_row}"])
        for offset, text in enumerate(validation_result["explanations"], start=1):
            sheet[f"A{expl_row + offset}"] = text


def _write_excel_sanity_checks(sheet, start: int, rows: list[dict[str, str]]) -> int:
    """Write categorized Excel-only sanity-check rows and return the next section row."""
    sheet[f"A{start}"] = "Sanity checks"
    _subheader(sheet[f"A{start}"])
    header_row = start + 1
    sheet[f"A{header_row}"], sheet[f"B{header_row}"], sheet[f"C{header_row}"] = "Severity", "Category", "Message"
    _header(sheet[f"A{header_row}:C{header_row}"])
    for idx, row in enumerate(rows, start=header_row + 1):
        sheet[f"A{idx}"] = row.get("severity")
        sheet[f"B{idx}"] = row.get("category")
        sheet[f"C{idx}"] = row.get("message")
        _style_severity_cell(sheet[f"A{idx}"], row.get("severity"))
        sheet[f"C{idx}"].alignment = Alignment(wrap_text=True, vertical="top")
    return header_row + len(rows) + 3


def _merge_excel_sanity_rows(runtime_rows: list[dict[str, str]], analyst_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    """Combine runtime gate warnings with analyst-level Excel checks."""
    rows = list(runtime_rows)
    if analyst_rows == [{"severity": "info", "category": "Overall", "message": "All sanity checks passed."}] and rows:
        return rows
    rows.extend(analyst_rows)
    if not rows:
        return [{"severity": "info", "category": "Overall", "message": "All sanity checks passed."}]
    return sorted(rows, key=lambda row: {"critical": 0, "warning_high": 1, "warning": 1, "info": 2}.get(row.get("severity"), 3))


def _style_severity_cell(cell, severity: str | None) -> None:
    """Apply existing severity color coding to a validation severity cell."""
    severity_text = str(severity or "").lower()
    cell.border = THIN_BORDER
    cell.font = Font(bold=True)
    if severity_text == "critical":
        cell.fill = PatternFill("solid", fgColor=STATUS_RED)
        cell.font = Font(bold=True, color=STATUS_RED_TEXT)
    elif severity_text == "warning":
        cell.fill = PatternFill("solid", fgColor=STATUS_YELLOW)
        cell.font = Font(bold=True, color=STATUS_YELLOW_TEXT)
    else:
        cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)


def _write_validation_dcf_tier_used(sheet, start_row: int, valuation: dict[str, Any]) -> None:
    """Write the DCF tier selection diagnostic to the Validation worksheet."""
    selected = valuation.get("selected_dcf_tier") or {}
    if not selected:
        return
    sheet[f"A{start_row}"] = "Valuation Tier Used"
    _subheader(sheet[f"A{start_row}"])
    sheet[f"A{start_row + 1}"], sheet[f"B{start_row + 1}"] = "Severity", "Message"
    _header(sheet[f"A{start_row + 1}:B{start_row + 1}"])
    severity = "info"
    if selected.get("tier") == 2:
        severity = "warning"
    elif selected.get("tier") == 3:
        severity = "warning"
    elif selected.get("tier") == 4:
        severity = "warning"
    elif selected.get("tier") == 5:
        severity = "critical"
    sheet[f"A{start_row + 2}"] = severity
    sheet[f"B{start_row + 2}"] = selected.get("selection_reason") or valuation.get("dcf_selection_reason")


def _build_raw_data_sheet(sheet, data: dict, income: pd.DataFrame, balance: pd.DataFrame, cash: pd.DataFrame, beta_match: Any) -> None:
    """
    Build the Raw Data worksheet.

    Formula: not applicable.
    Source: yfinance statements and matched Damodaran row.
    Example: stores sources, timestamps, raw statement tables, and Damodaran row values.
    Required inputs: data dict, statement frames, beta match.
    Limitation: compact raw dump, not a full yfinance object archive.
    """
    _title(sheet, "Raw Data")
    info = data.get("info", {})
    rows = [
        ("Source", DATA_SOURCES),
        ("Generated", current_timestamp()),
        ("Disclaimer", DISCLAIMER),
        ("Ticker", data.get("ticker")),
        ("financialCurrency", info.get("financialCurrency")),
        ("currentPrice", info.get("currentPrice") or info.get("regularMarketPrice")),
        ("cash", _safe_latest(balance, "cash")),
        ("sharesOutstanding", info.get("sharesOutstanding")),
        ("Damodaran source", beta_match.source_url),
        ("Damodaran industry", beta_match.matched_industry),
    ]
    for idx, row in enumerate(rows, start=3):
        sheet[f"A{idx}"], sheet[f"B{idx}"] = row
    _write_frame(sheet, "Income Statement Metrics", income, 15)
    _write_frame(sheet, "Balance Sheet Metrics", balance, 15 + len(income) + 4)
    _write_frame(sheet, "Cash Flow Metrics", cash, 15 + len(income) + len(balance) + 8)


def _write_frame(sheet, title: str, frame: pd.DataFrame, start_row: int) -> None:
    """
    Write a pandas DataFrame to a worksheet.

    Formula: not applicable.
    Source: internal workbook helper.
    Example: _write_frame(sheet, "Income", income, 15).
    Required inputs: sheet, title, frame, start row.
    Limitation: writes display values, not original raw yfinance statement objects.
    """
    sheet[f"A{start_row}"] = title
    _subheader(sheet[f"A{start_row}"])
    if frame.empty:
        sheet[f"A{start_row + 1}"] = "No data"
        return
    display, excluded_periods = _filter_empty_numeric_display_rows(frame)
    header_row = start_row + 1
    if excluded_periods:
        _write_exclusion_note(sheet, header_row, excluded_periods, max(len(display.columns), 1))
        header_row += 1
    for col_idx, col in enumerate(display.columns, start=1):
        sheet.cell(header_row, col_idx).value = col
    _header(sheet[f"A{header_row}:{get_column_letter(len(display.columns))}{header_row}"])
    for row_idx, (_, row) in enumerate(display.iterrows(), start=header_row + 1):
        for col_idx, value in enumerate(row.tolist(), start=1):
            sheet.cell(row_idx, col_idx).value = value


def _filter_empty_numeric_display_rows(frame: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    """Return a display-only DataFrame without rows whose numeric fields are all missing."""
    if frame.empty:
        return frame.copy(), []
    metadata_columns = {"year", "period", "period_end", "period_type"}
    data_columns = [column for column in frame.columns if column not in metadata_columns]
    if not data_columns:
        return frame.copy(), []
    numeric_values = frame[data_columns].apply(pd.to_numeric, errors="coerce")
    empty_mask = numeric_values.isna().all(axis=1)
    excluded = [_display_period_label(row) for _, row in frame.loc[empty_mask].iterrows()]
    return frame.loc[~empty_mask].copy(), excluded


def _write_exclusion_note(sheet, row_idx: int, excluded_periods: list[str], end_col: int) -> None:
    """Write an italicized note for historical rows hidden from the display layer."""
    note = _format_exclusion_note(excluded_periods)
    sheet.cell(row_idx, 1).value = note
    sheet.cell(row_idx, 1).font = Font(italic=True, color="666666")
    sheet.cell(row_idx, 1).alignment = Alignment(wrap_text=True, vertical="top")
    if end_col > 1:
        sheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=end_col)


def _format_exclusion_note(excluded_periods: list[str]) -> str:
    """Format the Yahoo Finance incomplete-data exclusion note."""
    if not excluded_periods:
        return ""
    ordered = sorted(excluded_periods, key=_period_sort_key)
    if len(ordered) == 1 or ordered[0] == ordered[-1]:
        period_text = ordered[0]
    else:
        period_text = f"{ordered[0]}-{ordered[-1]}"
    return f"Note: {period_text} excluded due to incomplete data from Yahoo Finance."


def _period_sort_key(label: str) -> int:
    """Sort FY labels by their year when possible."""
    digits = "".join(ch for ch in str(label) if ch.isdigit())
    return int(digits[:4]) if len(digits) >= 4 else 0


def _display_period_label(row: pd.Series) -> str:
    """Return a compact FY label for hidden historical rows."""
    if pd.notna(row.get("period")):
        return str(row.get("period")).split()[0]
    if pd.notna(row.get("year")):
        return f"FY{int(row.get('year'))}"
    return "earlier FY"


def _safe_latest(frame: pd.DataFrame, column: str) -> float | None:
    """
    Read the latest numeric value from a DataFrame column.

    Formula: not applicable.
    Source: internal helper.
    Example: latest revenue from income metrics.
    Required inputs: frame and column.
    Limitation: returns None when no numeric value exists.
    """
    if frame.empty or column not in frame.columns:
        return None
    values = pd.to_numeric(frame[column], errors="coerce").dropna()
    return float(values.iloc[-1]) if not values.empty else None


def _latest_period(frame: pd.DataFrame) -> str:
    """Return the latest fiscal period label for Excel assumption sources."""
    if frame.empty:
        return "latest FY"
    row = frame.iloc[-1]
    if "period" in frame.columns and pd.notna(row.get("period")):
        return str(row.get("period")).split(" ")[0]
    if "year" in frame.columns and pd.notna(row.get("year")):
        return f"FY{int(row.get('year'))}"
    return "latest FY"


def _period_range(frame: pd.DataFrame) -> str:
    """Return a compact fiscal period range for Excel source notes."""
    if frame.empty:
        return "available FY history"
    labels = []
    for _, row in frame.iterrows():
        if "period" in frame.columns and pd.notna(row.get("period")):
            labels.append(str(row.get("period")).split(" ")[0])
        elif "year" in frame.columns and pd.notna(row.get("year")):
            labels.append(f"FY{int(row.get('year'))}")
    if not labels:
        return "available FY history"
    if len(labels) == 1:
        return labels[0]
    return f"{labels[0]}-{labels[-1]}"


def _implied_price_source_note(raw_implied_price: float | None, currency: str) -> str:
    """Return the DCF source note for the visible clamped implied share price."""
    if raw_implied_price is None:
        return "Calculated DCF result"
    try:
        raw_value = float(raw_implied_price)
    except (TypeError, ValueError):
        return "Calculated DCF result"
    if raw_value < 0:
        return (
            f"DCF model produced negative equity value ({raw_value:.2f} {currency} raw); clamped to zero. "
            "This indicates assumptions imply firm worthless under current trajectory; review Scenario tab"
        )
    return "Calculated DCF result"




def _autofit(sheet) -> None:
    """
    Auto-size worksheet columns with a reasonable cap.

    Formula: width based on max displayed text length.
    Source: internal workbook formatting.
    Example: called for every worksheet before export.
    Required inputs: worksheet.
    Limitation: approximate width, not Excel's native autofit.
    """
    for column in sheet.columns:
        max_length = 8
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.value is not None:
                max_length = max(max_length, min(len(str(cell.value)), 60))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.column_dimensions[column_letter].width = min(max_length + 2, 42)
