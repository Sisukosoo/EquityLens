"""Tests for valuation Excel workbook generation."""

from io import BytesIO
from types import SimpleNamespace

import pandas as pd
import pytest
from openpyxl import load_workbook

from utils.excel_report import _find_libreoffice_executable, _tier_status_text, build_valuation_excel_report


def _sheet_values(sheet) -> set:
    return {cell.value for row in sheet.iter_rows() for cell in row if cell.value is not None}


def _summary_value(sheet, label: str):
    for row in range(1, sheet.max_row + 1):
        if sheet[f"A{row}"].value == label:
            return sheet[f"B{row}"].value
    raise AssertionError(f"Summary label not found: {label}")


def _summary_tier_status(sheet, method: str):
    for row in range(1, sheet.max_row + 1):
        if sheet[f"A{row}"].value == method:
            return sheet[f"D{row}"].value
    raise AssertionError(f"Summary tier row not found: {method}")


def _tier(tier: int, selected: bool = False) -> dict:
    dcf = {
        "forecast": [
            {"year": 1, "revenue": 100.0, "ebit": 10.0, "fcf": 7.0, "pv_fcf": 6.5},
            {"year": 2, "revenue": 102.5, "ebit": 10.3, "fcf": 7.2, "pv_fcf": 6.2},
        ],
        "terminal_value": 130.0,
        "pv_terminal_value": 90.0,
        "enterprise_value": 102.7,
        "equity_value": 100.0,
        "implied_price": 10.0 + tier,
        "upside": 0.05,
    }
    return {
        "tier": tier,
        "name": ["Standard DCF", "Smoothed DCF", "Sector Benchmark DCF"][tier - 1],
        "method": f"Tier {tier}",
        "confidence": "LOW" if tier == 3 else "NORMAL",
        "selected": selected,
        "status": "ACCEPTED" if selected else "REJECTED",
        "rejection_reason": "" if selected else "failed sanity check",
        "acceptance_reason": "within 70% of market price" if selected else "",
        "selection_reason": "selected for test" if selected else "",
        "explanation": "test explanation",
        "assumptions": {
            "revenue_growth": 0.025,
            "revenue_growth_source": "test",
            "ebit_margin": 0.10,
            "ebit_margin_source": "test",
            "depreciation_pct_revenue": 0.03,
            "depreciation_source": "test",
            "capex_pct_revenue": 0.04,
            "capex_source": "test",
            "working_capital_pct_revenue": 0.02,
            "working_capital_source": "test",
        },
        "dcf": dcf,
    }


def _build_sample_workbook_bytes():
    income = pd.DataFrame(
        {
            "year": [2023, 2024],
            "period": ["FY2023", "FY2024"],
            "revenue": [95.0, 100.0],
            "ebit": [9.0, 10.0],
            "ebit_margin": [9.5, 10.0],
        }
    )
    balance = pd.DataFrame(
        {
            "year": [2024],
            "cash": [5.0],
            "current_assets": [30.0],
            "current_liabilities": [20.0],
        }
    )
    cash = pd.DataFrame({"year": [2023, 2024], "free_cash_flow": [6.0, 7.0]})

    return _build_sample_workbook_bytes_with_statements(income, balance, cash)


def _build_sample_workbook_bytes_with_statements(
    income: pd.DataFrame,
    balance: pd.DataFrame,
    cash: pd.DataFrame,
    sanity_warnings: list[dict] | None = None,
):
    tiers = [_tier(1), _tier(2), _tier(3)]
    tier4 = {
        "tier": 4,
        "name": "Multiples-Based Valuation",
        "method": "Tier 4 - multiples valuation",
        "confidence": "LOW",
        "selected": True,
        "status": "ACCEPTED",
        "rejection_reason": "",
        "acceptance_reason": "within 70% of market price",
        "selection_reason": "selected for test",
        "explanation": "multiples selected",
        "assumptions": {"source": "Sector multiples (median of EV/EBITDA, EV/Sales, P/Book)"},
        "multiples": [
            {"method": "EV/EBITDA", "multiple": 5.5, "implied_price": 11.0, "source": "Damodaran sector benchmark: Test"},
            {"method": "EV/Sales", "multiple": 0.9, "implied_price": 12.0, "source": "Damodaran sector benchmark: Test"},
            {"method": "P/Book", "multiple": 1.3, "implied_price": 13.0, "source": "Damodaran sector benchmark: Test"},
        ],
        "dcf": {"implied_price": 12.0, "upside": 0.0},
    }
    tier5 = {
        "tier": 5,
        "name": "Tangible Book Value Floor",
        "method": "Tier 5 - tangible book floor",
        "confidence": "VERY LOW",
        "selected": False,
        "status": "REFERENCE",
        "rejection_reason": "",
        "assumptions": {"total_equity": 9.5, "shares_outstanding": 1_000_000},
        "dcf": {"implied_price": 9.5, "upside": -0.2},
    }
    tiers.extend([tier4, tier5])
    valuation = {
        "currency": "EUR",
        "market_cap": 1000.0,
        "total_debt": 200.0,
        "de_ratio": 0.2,
        "unlevered_beta": 0.9,
        "levered_beta": 1.035,
        "yfinance_beta": 1.2,
        "risk_free_rate": 0.04,
        "risk_free_date": "2026-01-31",
        "market_risk_premium": 0.055,
        "cost_of_equity": 0.096925,
        "cost_of_debt": 0.04,
        "cost_of_debt_estimated": False,
        "equity_weight": 1000.0 / 1200.0,
        "debt_weight": 200.0 / 1200.0,
        "wacc": 0.08577083333333334,
        "tax_rate": 0.25,
        "terminal_growth": 0.025,
        "current_price": 12.0,
        "dcf": tier4["dcf"],
        "dcf_tiers": tiers,
        "selected_dcf_tier": tier4,
        "reverse_dcf": {
            "implied_growth": 0.06,
            "tier1_growth": 0.025,
            "tier1_growth_source": "1-year growth rate (FY2023-FY2024)",
            "analyst_consensus_growth": None,
            "analyst_consensus_source": "yfinance revenueGrowth",
            "growth_gap": 0.035,
            "interpretation": "Market is pricing in higher growth than the model assumes - this explains the negative valuation gap.",
            "source": "Reverse DCF (solved from market price)",
        },
    }
    data = {
        "ticker": "TEST",
        "info": {
            "longName": "Test Company",
            "industry": "Test Industry",
            "financialCurrency": "EUR",
            "currentPrice": 12.0,
            "sharesOutstanding": 1_000_000,
        },
    }
    beta_match = SimpleNamespace(
        source_url="https://example.test/betas.xls",
        source_updated="January 2026",
        company_industry="Test Industry",
        matched_industry="Test Industry",
        confidence=95.0,
    )

    return build_valuation_excel_report(data, income, balance, cash, beta_match, valuation, None, sanity_warnings or [])


def _build_sample_workbook():
    return load_workbook(BytesIO(_build_sample_workbook_bytes()), data_only=False)


def test_excel_report_includes_dcf_tier_sections():
    workbook = _build_sample_workbook()

    assert "DCF" in workbook.sheetnames
    assert "Implied Share Price Analysis" in _sheet_values(workbook["Summary"])
    assert "Tier 1 Assumptions - Standard DCF" in _sheet_values(workbook["DCF"])
    assert "Tier 4 - Multiples-Based Valuation" in _sheet_values(workbook["DCF"])
    assert "Tier 5 - Tangible Book Floor" in _sheet_values(workbook["DCF"])
    assert "Final Selected Result" in _sheet_values(workbook["DCF"])
    assert "Reverse DCF Analysis" in _sheet_values(workbook["DCF"])
    assert "Valuation Tier Used" in _sheet_values(workbook["Validation"])


def test_excel_report_beta_capm_wacc_formula_cells_match_expected_values():
    workbook = _build_sample_workbook()
    beta = workbook["Beta (Damodaran)"]
    capm = workbook["CAPM"]
    wacc = workbook["WACC"]

    assert beta["B11"].value == "=B8*(1+(1-B10)*B9)"
    expected_beta = beta["B8"].value * (1 + (1 - beta["B10"].value) * beta["B9"].value)
    assert abs(expected_beta - 1.035) < 0.000001

    assert capm["B4"].value == "='Beta (Damodaran)'!B11"
    assert capm["B6"].value == "=B3+B4*B5"
    expected_cost_of_equity = capm["B3"].value + expected_beta * capm["B5"].value
    assert abs(expected_cost_of_equity - 0.096925) < 0.000001

    assert wacc["B5"].value == "=B3+B4"
    assert wacc["B6"].value == "=B3/B5"
    assert wacc["B7"].value == "=B4/B5"
    assert wacc["B8"].value == "='CAPM'!B6"
    assert wacc["B11"].value == "=B9*(1-B10)"
    assert wacc["B12"].value == "=(B6*B8)+(B7*B9*(1-B10))"
    expected_wacc = (1000.0 / 1200.0) * expected_cost_of_equity + (200.0 / 1200.0) * 0.04 * (1 - 0.25)
    assert abs(expected_wacc - 0.08577083333333334) < 0.000001


def test_excel_report_reverse_dcf_sources_are_row_specific():
    workbook = _build_sample_workbook()
    sheet = workbook["DCF"]
    start_row = next(row for row in range(1, sheet.max_row + 1) if sheet[f"A{row}"].value == "Reverse DCF Analysis")
    sources = [sheet[f"C{row}"].value for row in range(start_row + 2, start_row + 7)]

    assert sources == [
        "Solved from current market price using Tier 1 DCF inputs",
        "Same as Tier 1 DCF assumption: 1-year growth rate (FY2023-FY2024)",
        "yfinance revenueGrowth",
        "Calculated: implied minus Tier 1 assumed (in percentage points)",
        None,
    ]
    assert len(set(sources)) == 5


def test_excel_report_filters_empty_historical_display_rows_and_notes_exclusion():
    income = pd.DataFrame(
        {
            "year": [2021, 2022],
            "period": ["FY2021", "FY2022"],
            "revenue": [None, 100.0],
            "ebit": [None, 10.0],
            "ebit_margin": [None, 10.0],
        }
    )
    balance = pd.DataFrame(
        {
            "year": [2021, 2022],
            "period": ["FY2021", "FY2022"],
            "cash": [None, 20.0],
            "current_assets": [None, 30.0],
        }
    )
    cash = pd.DataFrame(
        {
            "year": [2021, 2022],
            "period": ["FY2021", "FY2022"],
            "free_cash_flow": [None, 7.0],
        }
    )
    workbook = load_workbook(
        BytesIO(_build_sample_workbook_bytes_with_statements(income, balance, cash)),
        data_only=False,
    )

    dcf_values = _sheet_values(workbook["DCF"])
    raw_values = _sheet_values(workbook["Raw Data"])
    dcf_period_values = [workbook["DCF"][f"A{row}"].value for row in range(1, workbook["DCF"].max_row + 1)]

    assert "Note: FY2021 excluded due to incomplete data from Yahoo Finance." in dcf_values
    assert "Note: FY2021 excluded due to incomplete data from Yahoo Finance." in raw_values
    assert "FY2021" not in dcf_period_values


def test_excel_report_adds_yfinance_beta_context_note():
    workbook = _build_sample_workbook()

    assert workbook["Beta (Damodaran)"]["C12"].value.startswith(
        "Yahoo Finance beta typically reflects 1-2 year regression against the market"
    )


def test_excel_report_recalc_populates_cached_formula_values_when_libreoffice_available():
    if _find_libreoffice_executable() is None:
        pytest.skip("LibreOffice is not installed; cached formula recalc test skipped.")

    workbook = load_workbook(BytesIO(_build_sample_workbook_bytes()), data_only=True)

    assert isinstance(workbook["Beta (Damodaran)"]["B11"].value, (int, float))
    assert isinstance(workbook["CAPM"]["B6"].value, (int, float))
    assert isinstance(workbook["WACC"]["B12"].value, (int, float))


def test_excel_report_accepted_tier_status_includes_reason():
    status = _tier_status_text({"status": "ACCEPTED", "acceptance_reason": "within 70% of market price"})

    assert status == "ACCEPTED - within 70% of market price"
    assert status.strip() != "ACCEPTED -"


def test_excel_report_tier5_reference_status_has_clear_reason():
    status = _tier_status_text({"tier": 5, "status": "REFERENCE", "rejection_reason": ""})

    assert status == "REFERENCE - tangible book value shown as conservative floor"
    assert "no reason provided" not in status


def test_excel_report_includes_runtime_business_model_sanity_warning():
    income = pd.DataFrame({"year": [2024], "period": ["FY2024"], "revenue": [100.0], "ebit_margin": [10.0]})
    balance = pd.DataFrame({"year": [2024], "period": ["FY2024"], "cash": [20.0]})
    cash = pd.DataFrame({"year": [2024], "period": ["FY2024"], "free_cash_flow": [7.0]})
    message = "DCF model is not appropriate for Bank (Money Center) businesses."
    workbook = load_workbook(
        BytesIO(
            _build_sample_workbook_bytes_with_statements(
                income,
                balance,
                cash,
                [
                    {
                        "severity": "critical",
                        "category": "Business Model Compatibility",
                        "message": message,
                    }
                ],
            )
        ),
        data_only=False,
    )
    sheet = workbook["Validation"]
    sanity_row = next(row for row in range(1, sheet.max_row + 1) if sheet[f"A{row}"].value == "Sanity checks")

    assert sheet[f"A{sanity_row + 2}"].value == "critical"
    assert sheet[f"B{sanity_row + 2}"].value == "Business Model Compatibility"
    assert sheet[f"C{sanity_row + 2}"].value == message


def test_summary_tab_reflects_critical_sanity_warning():
    workbook = load_workbook(
        BytesIO(
            _build_sample_workbook_bytes_with_statements(
                pd.DataFrame({"year": [2024], "period": ["FY2024"], "revenue": [100.0], "ebit_margin": [10.0]}),
                pd.DataFrame({"year": [2024], "period": ["FY2024"], "cash": [20.0]}),
                pd.DataFrame({"year": [2024], "period": ["FY2024"], "free_cash_flow": [7.0]}),
                [
                    {
                        "severity": "critical",
                        "category": "Business Model Compatibility",
                        "message": "DCF model is not appropriate for Bank (Money Center) businesses.",
                    }
                ],
            )
        ),
        data_only=False,
    )
    sheet = workbook["Summary"]

    assert sheet["A1"].value.startswith("⚠ CRITICAL WARNING")
    assert sheet["A2"].value == "Valuation Summary"
    assert _summary_value(sheet, "Confidence") == (
        "CRITICAL - DCF model not appropriate for this business type. See Validation tab. "
        "Result should not be used as a fair value estimate."
    )
    assert _summary_value(sheet, "Reason") == (
        "One or more critical sanity checks have fired. See Validation tab > Sanity Checks for details."
    )
    assert _summary_tier_status(sheet, "Tier 4 - Multiples-Based Valuation") == (
        "Computed but flagged - critical sanity check applies. See Validation tab."
    )


def test_summary_tab_stays_normal_without_critical_sanity_warning():
    workbook = _build_sample_workbook()
    sheet = workbook["Summary"]

    assert sheet["A1"].value == "Valuation Summary"
    assert _summary_value(sheet, "Confidence") == "LOW - multiples-based valuation; DCF model not applicable"
    assert _summary_value(sheet, "Reason") == "multiples selected"
    assert _summary_tier_status(sheet, "Tier 4 - Multiples-Based Valuation") == (
        "Used as primary result - within 70% of market price"
    )
